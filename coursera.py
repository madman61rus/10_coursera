import requests
import json
from lxml import etree
import random
from bs4 import BeautifulSoup
from openpyxl import Workbook
import argparse


def get_courses_random_list(url, number_results):
    courses_list = requests.get(url)
    courses_xml = etree.fromstring(courses_list.content)
    courses_urls = [child[0].text for child in courses_xml]
    return random.sample(courses_urls, number_results)


def get_course_info(url):
    try:
        course_html = requests.get(url)
        soup = BeautifulSoup(course_html.content, 'html.parser')
        result_info = {
            'course_url' : url,
            'title' : get_courses_title(soup),
            'language' : get_courses_language(soup),
            'start_date' : get_courses_start(soup),
            'weeks' : get_courses_weeks(soup),
            'stars' : get_courses_stars(soup)
        }
        return (result_info)
    except:
        return None


def get_courses_title(soup):
    return soup.find('div', {'class': 'title'}).contents[0]


def get_courses_language(soup):
    return soup.find('div', {'class': 'language-info'}).contents[1]


def get_courses_start(soup):
    data_from_script = soup.select('script[type="application/ld+json"]')[0].text
    data_json = json.loads(data_from_script)
    start_date = data_json['hasCourseInstance'][0]['startDate']
    if start_date:
        return start_date
    else:
        return None


def get_courses_weeks(soup):
    return len(soup.find_all('div', {'class': 'week-heading body-2-text'}))


def get_courses_stars(soup):
    stars = soup.find('div', {'class': 'ratings-text bt3-visible-xs'}).contents[0]
    return stars.replace('stars', '')

def prepare_output_courses(cources):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Title'
    ws['B1'] = 'Language'
    ws['C1'] = 'Start Date'
    ws['D1'] = 'Weeks'
    ws['E1'] = 'Stars'
    ws['F1'] = 'Url'

    for counter, course_info in enumerate(cources):
        counter += 2
        ws.cell(row=counter, column=1, value=course_info['title'])
        ws.cell(row=counter, column=2, value=course_info['language'])
        ws.cell(row=counter, column=3, value=course_info['start_date'])
        ws.cell(row=counter, column=4, value=course_info['weeks'])
        ws.cell(row=counter, column=5, value=course_info['stars'])
        ws.cell(row=counter, column=6, value=course_info['course_url'])

    return wb

def output_courses_info_to_xlsx(filepath, workbook):

    workbook.save(filename=filepath)


if __name__ == '__main__':
    args = argparse.ArgumentParser()
    args.add_argument('-f', '--file', help='Файл для сохранения результатов', required=True)
    args.add_argument('-r', '--results', help='Количество выводимых результатов')
    arguments = args.parse_args()

    print('Получаем информацию с coursera.org...')
    courses_urls = get_courses_random_list('https://www.coursera.org/sitemap~www~courses.xml', int(arguments.results))
    print('Обрабатываем информацию ...')
    courses_info = [get_course_info(url) for url in courses_urls if get_course_info(url)]
    print('Сохраняем информацию в файл')
    output_courses_info_to_xlsx(arguments.file, prepare_output_courses(courses_info))
